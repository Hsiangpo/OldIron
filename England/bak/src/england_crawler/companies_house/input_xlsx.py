"""英国 xlsx 公司名单读取。"""

from __future__ import annotations

from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

from england_crawler.companies_house.client import normalize_company_name


HEADER_NAMES = {"COMPANYNAME", "COMPANY NAME"}


def iter_company_names_from_xlsx(path: str | Path, limit: int = 0) -> Iterator[str]:
    """按顺序读取公司名，自动跳过表头、空行与标准化重复项。"""
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        seen: set[str] = set()
        yielded = 0
        for row in sheet.iter_rows(values_only=True):
            raw_name = str((row[0] if row else "") or "").strip()
            if not raw_name:
                continue
            normalized = normalize_company_name(raw_name)
            if not normalized or normalized in HEADER_NAMES or normalized in seen:
                continue
            seen.add(normalized)
            yield raw_name
            yielded += 1
            if limit > 0 and yielded >= limit:
                return
    finally:
        workbook.close()


def load_company_names_from_xlsx(path: str | Path, limit: int = 0) -> list[str]:
    """一次性加载公司名列表。"""
    return list(iter_company_names_from_xlsx(path, limit=limit))
