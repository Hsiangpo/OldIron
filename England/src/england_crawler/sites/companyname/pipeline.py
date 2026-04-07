"""CompanyName 主流程：公司名 → GMap 查官网 → 邮箱补充。"""

from __future__ import annotations

import logging
import re
import threading
import time
from pathlib import Path

from england_crawler.sites.companyname.companies_house import CompaniesHouseClient
from england_crawler.sites.companyname.email_rules import EnglandRuleEmailExtractor
from oldiron_core.fc_email.domain_cache import FirecrawlDomainCache
from oldiron_core.fc_email.email_service import FirecrawlEmailService, FirecrawlEmailSettings
from oldiron_core.google_maps import GoogleMapsClient, GoogleMapsConfig
from england_crawler.sites.companyname.config import CompanyNameConfig
from england_crawler.sites.companyname.store import (
    CompanyNameStore, FirecrawlTask, GMapTask,
)

try:
    from oldiron_core.protocol_crawler import SiteCrawlClient, SiteCrawlConfig
except ImportError:
    SiteCrawlClient = None  # type: ignore[assignment,misc]
    SiteCrawlConfig = None  # type: ignore[assignment,misc]


LOGGER = logging.getLogger(__name__)


def _retry_delay_seconds(retries: int, cap_seconds: float) -> float:
    return min(float(2 ** max(int(retries), 0)), float(cap_seconds))


def _build_gmap_queries(task: GMapTask) -> list[str]:
    """为英国公司构建 GMap 查询列表。"""
    queries: list[str] = []
    for value in (
        f"{task.company_name} England",
        f"{task.company_name} United Kingdom",
        task.company_name,
    ):
        text = re.sub(r"\s+", " ", str(value or "").strip())
        if text and text not in queries:
            queries.append(text)
    return queries


class CompanyNamePipelineRunner:
    """协调 GMap → 官网爬虫(邮箱补充)。"""

    def __init__(
        self,
        *,
        config: CompanyNameConfig,
        skip_gmap: bool,
        skip_firecrawl: bool,
    ) -> None:
        self.config = config
        self.skip_gmap = skip_gmap
        self.skip_firecrawl = skip_firecrawl
        self.store = CompanyNameStore(config.store_db_path)
        self.stop_event = threading.Event()
        self._gmap_local = threading.local()
        self._firecrawl_local = threading.local()
        self._companies_house_local = threading.local()
        self._firecrawl_key_pool = None
        self._firecrawl_key_pool_lock = threading.Lock()
        self._firecrawl_domain_cache = FirecrawlDomainCache(
            self.config.project_root / "output" / "firecrawl_cache.db"
        )
        self._firecrawl_settings = FirecrawlEmailSettings(
            keys_inline=list(self.config.firecrawl_keys_inline or []),
            keys_file=Path(self.config.firecrawl_keys_file or (self.config.project_root / "output" / "firecrawl_keys.txt")),
            pool_db=Path(self.config.firecrawl_pool_db or (self.config.project_root / "output" / "cache" / "firecrawl_keys.db")),
            base_url=self.config.firecrawl_base_url,
            timeout_seconds=self.config.firecrawl_timeout_seconds,
            max_retries=self.config.firecrawl_max_retries,
            key_per_limit=self.config.firecrawl_key_per_limit,
            key_wait_seconds=self.config.firecrawl_key_wait_seconds,
            key_cooldown_seconds=self.config.firecrawl_key_cooldown_seconds,
            key_failure_threshold=self.config.firecrawl_key_failure_threshold,
            llm_api_key=self.config.llm_api_key,
            llm_base_url=self.config.llm_base_url,
            llm_model=self.config.llm_model,
            llm_reasoning_effort=self.config.llm_reasoning_effort,
            llm_api_style=self.config.llm_api_style,
            llm_timeout_seconds=self.config.llm_timeout_seconds,
            prefilter_limit=self.config.firecrawl_prefilter_limit,
            llm_pick_count=self.config.firecrawl_llm_pick_count,
            extract_max_urls=self.config.firecrawl_extract_max_urls,
            zero_retry_seconds=self.config.firecrawl_zero_retry_seconds,
            contact_form_retry_seconds=self.config.firecrawl_contact_form_retry_seconds,
            crawl_backend=self.config.crawl_backend,
        )

    def run(self) -> None:
        self.config.validate(skip_firecrawl=self.skip_firecrawl)
        self._seed_from_excel()
        recovered = self.store.requeue_stale_running_tasks(
            older_than_seconds=self.config.stale_running_requeue_seconds
        )
        if recovered:
            LOGGER.info("已回收陈旧运行中任务：%s", recovered)
        # 启动前批量处理：域名缓存已有结果的 pending 任务直接标 done
        cached = self._firecrawl_domain_cache.get_all_done_domains()
        batch_resolved = self.store.batch_resolve_cached_firecrawl(cached)
        if batch_resolved:
            LOGGER.info("启动预处理：批量跳过 %d 个已缓存域名的任务", batch_resolved)
        workers = self._build_workers()
        for w in workers:
            w.start()
            # email worker 错峰启动，避免同时请求 LLM API
            if w.name.startswith("email-"):
                time.sleep(0.1)
        try:
            self._monitor_until_done()
        finally:
            self.stop_event.set()
            for w in workers:
                w.join(timeout=2)
            self.store.export_jsonl_snapshots(self.config.output_dir)
            self._firecrawl_domain_cache.close()
            self.store.close()

    def _seed_from_excel(self) -> None:
        """从 Excel 文件灌入公司名（数据库已有则跳过）。"""
        existing = self.store.company_count()
        if existing > 0 and not getattr(self.config, "reseed", False):
            LOGGER.info("数据库已有 %d 家公司，跳过 Excel 读取（使用 --reseed 强制重读）", existing)
            return

        import openpyxl

        all_names: list[str] = []
        seen: set[str] = set()
        for xlsx_path in self.config.excel_files:
            if not xlsx_path.exists():
                LOGGER.warning("Excel 文件不存在：%s", xlsx_path)
                continue
            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        name = str(cell or "").strip()
                        if not name or name.lower() in ("companyname", "company_name", "company name"):
                            continue
                        key = name.lower()
                        if key not in seen:
                            seen.add(key)
                            all_names.append(name)
            wb.close()
            LOGGER.info("读取 Excel：%s | 累计去重公司名 %d", xlsx_path.name, len(all_names))

        if not all_names:
            LOGGER.warning("没有读取到任何公司名")
            return

        added = self.store.seed_companies(all_names)
        LOGGER.info("公司名灌入完成：新增 %d / 总去重 %d", added, len(all_names))

    def _build_workers(self) -> list[threading.Thread]:
        workers: list[threading.Thread] = []
        if not self.skip_gmap:
            workers.extend(
                threading.Thread(target=self._gmap_worker, name=f"gmap-{i+1}", daemon=True)
                for i in range(self.config.gmap_workers)
            )
        if not self.skip_firecrawl:
            workers.extend(
                threading.Thread(target=self._email_worker, name=f"email-{i+1}", daemon=True)
                for i in range(self.config.firecrawl_workers)
            )
        return workers

    def _monitor_until_done(self) -> None:
        last_log = 0.0
        last_snapshot = 0.0
        while not self.stop_event.is_set():
            now = time.monotonic()
            self.store.requeue_stale_running_tasks(
                older_than_seconds=self.config.stale_running_requeue_seconds
            )
            if now - last_log >= self.config.log_interval_seconds:
                p = self.store.progress()
                LOGGER.info(
                    "进度：gmap=%d/%d email=%d/%d companies=%d final=%d",
                    p.gmap_running, p.gmap_pending,
                    p.firecrawl_running, p.firecrawl_pending,
                    p.companies_total, p.final_total,
                )
                last_log = now
                # 全部完成检查
                if p.gmap_pending == 0 and p.gmap_running == 0 and \
                   p.firecrawl_pending == 0 and p.firecrawl_running == 0:
                    LOGGER.info("所有任务完成")
                    break
            if now - last_snapshot >= self.config.snapshot_interval_seconds:
                self.store.export_jsonl_snapshots(self.config.output_dir)
                last_snapshot = now
            time.sleep(self.config.queue_poll_interval)

    # ── GMap Worker ──

    def _gmap_worker(self) -> None:
        client = self._get_gmap_client()
        while not self.stop_event.is_set():
            task = self.store.claim_gmap_task()
            if not task:
                time.sleep(1.0)
                continue
            try:
                self._do_gmap(task, client)
            except Exception as exc:  # noqa: BLE001
                delay = _retry_delay_seconds(task.retries, 60.0)
                LOGGER.warning("GMap 异常 %s：%s，%0.fs 后重试", task.orgnr, exc, delay)
                self.store.defer_gmap_task(task.orgnr, delay, str(exc)[:200])

    def _do_gmap(self, task: GMapTask, client: GoogleMapsClient) -> None:
        queries = _build_gmap_queries(task)
        LOGGER.info("GMap 开始：%s | %s", task.orgnr, task.company_name)
        for query in queries:
            LOGGER.info("GMap 查询：%s | %s", task.orgnr, query)
            result = client.search_company_profile(query, task.company_name)
            if result and result.website:
                LOGGER.info("GMap 完成：%s | 官网=%s", task.orgnr, result.website)
                self.store.complete_gmap_task(
                    task.orgnr,
                    homepage=result.website,
                    phone=result.phone or task.proff_phone,
                    gmap_name=result.company_name or "",
                    evidence_url=result.website,
                )
                return
        # 没找到
        LOGGER.info("GMap 完成：%s | 官网=-", task.orgnr)
        self.store.complete_gmap_task(
            task.orgnr, homepage="", phone=task.proff_phone,
            gmap_name="", evidence_url="",
        )

    def _get_gmap_client(self) -> GoogleMapsClient:
        client = getattr(self._gmap_local, "client", None)
        if client is None:
            client = GoogleMapsClient(GoogleMapsConfig(
                proxy_url=self.config.gmap_proxy,
            ))
            self._gmap_local.client = client
        return client

    # ── 协议规则邮箱发现 Worker ──

    def _email_worker(self) -> None:
        while not self.stop_event.is_set():
            task = self.store.claim_firecrawl_task()
            if not task:
                time.sleep(1.0)
                continue
            try:
                self._process_firecrawl_task(task)
            except Exception as exc:  # noqa: BLE001
                self._handle_firecrawl_task_exception(task, exc)

    def _handle_firecrawl_task_exception(self, task: FirecrawlTask, exc: Exception) -> None:
        attempt = task.retries + 1
        delay = _retry_delay_seconds(attempt, 120.0)
        LOGGER.exception(
            "邮箱补充 线程异常：%s | 域名=%s | 第%d次，%0.fs 后重试",
            task.orgnr,
            task.domain or "-",
            attempt,
            delay,
        )
        self._reset_email_worker_state()
        self.store.defer_firecrawl_task(task.orgnr, delay, str(exc)[:200])

    def _process_firecrawl_task(self, task: FirecrawlTask) -> None:
        ch_result = self._get_companies_house_client().lookup_company(task.company_name)
        LOGGER.info(
            "CH officers：%s | 编号=%s | 名字=%d | representative=%s",
            task.orgnr,
            ch_result.company_number or "-",
            len(ch_result.officer_names),
            ch_result.representative or "-",
        )
        if ch_result.company_number or ch_result.officer_names or ch_result.representative:
            self.store.update_companies_house_result(
                task.orgnr,
                company_number=ch_result.company_number,
                officers_url=ch_result.officers_url,
                officer_names=ch_result.officer_names,
                representative=ch_result.representative,
            )
        if not task.website:
            self.store.complete_firecrawl_task(
                task.orgnr,
                [],
                evidence_url=ch_result.officers_url or "",
                representative=ch_result.representative,
                company_number=ch_result.company_number,
                officers_url=ch_result.officers_url,
                officer_names=ch_result.officer_names,
            )
            return
        decision = self._firecrawl_domain_cache.prepare_lookup(
            task.domain, stale_running_seconds=30.0,
        )
        if decision.status == "done":
            if decision.emails:
                LOGGER.info("邮箱补充 命中缓存：%s | 域名=%s | 邮箱=%d", task.orgnr, task.domain, len(decision.emails))
                self.store.complete_firecrawl_task(
                    task.orgnr,
                    decision.emails,
                    evidence_url=ch_result.officers_url or task.website,
                    representative=ch_result.representative,
                    company_number=ch_result.company_number,
                    officers_url=ch_result.officers_url,
                    officer_names=ch_result.officer_names,
                )
            else:
                LOGGER.info("邮箱补充 命中缓存(无邮箱)：%s | 域名=%s", task.orgnr, task.domain)
                self.store.complete_firecrawl_task(
                    task.orgnr,
                    [],
                    evidence_url=ch_result.officers_url or task.website,
                    representative=ch_result.representative,
                    company_number=ch_result.company_number,
                    officers_url=ch_result.officers_url,
                    officer_names=ch_result.officer_names,
                )
            return
        if decision.status == "wait":
            LOGGER.info("邮箱补充 等待同域名：%s | 域名=%s", task.orgnr, task.domain)
            self.store.defer_firecrawl_task(task.orgnr, max(decision.wait_seconds, 1.0))
            return
        LOGGER.info("邮箱补充 开始：%s | 域名=%s", task.orgnr, task.domain)
        try:
            result = self._get_rule_email_extractor().discover(
                company_name=task.company_name,
                homepage=task.website,
                domain=task.domain,
            )
            emails = list(result.emails or [])
            ev = str(result.evidence_url or task.website).strip()
            LOGGER.info("邮箱补充 完成：%s | 域名=%s | 邮箱=%d | 代表人=%s", task.orgnr, task.domain, len(emails), ch_result.representative or "-")
            self._firecrawl_domain_cache.mark_done(task.domain, emails, retry_after_seconds=0.0)
            self.store.complete_firecrawl_task(
                task.orgnr,
                emails,
                evidence_url=ch_result.officers_url or ev,
                representative=ch_result.representative,
                company_number=ch_result.company_number,
                officers_url=ch_result.officers_url,
                officer_names=ch_result.officer_names,
            )
        except Exception as exc:  # noqa: BLE001
            attempt = task.retries + 1
            delay = _retry_delay_seconds(attempt, 120.0)
            if attempt >= 5:
                LOGGER.warning("邮箱补充 最终失败 %s：%s", task.orgnr, exc)
                self._firecrawl_domain_cache.mark_done(task.domain, [])
                self.store.complete_firecrawl_task(
                    task.orgnr,
                    [],
                    evidence_url=ch_result.officers_url or "",
                    representative=ch_result.representative,
                    company_number=ch_result.company_number,
                    officers_url=ch_result.officers_url,
                    officer_names=ch_result.officer_names,
                )
                return
            LOGGER.warning("邮箱补充 重试 %s(%d)：%s", task.orgnr, attempt, exc)
            self._firecrawl_domain_cache.defer(task.domain, delay_seconds=delay, error_text=str(exc)[:200])
            self.store.defer_firecrawl_task(task.orgnr, delay, str(exc)[:200])

    def _get_firecrawl_service(self) -> FirecrawlEmailService:
        svc = getattr(self._firecrawl_local, "service", None)
        if svc is None:
            firecrawl_client = None
            # 协议爬虫后端
            if self.config.crawl_backend == "protocol" and SiteCrawlClient is not None:
                LOGGER.info("使用协议爬虫后端 (CRAWL_BACKEND=protocol)")
                firecrawl_client = SiteCrawlClient(SiteCrawlConfig(
                    timeout_seconds=self.config.firecrawl_timeout_seconds,
                    max_retries=self.config.firecrawl_max_retries,
                ))
            key_pool = None if (self.config.crawl_backend == "protocol") else FirecrawlEmailService.build_key_pool(self._firecrawl_settings)
            svc = FirecrawlEmailService(
                self._firecrawl_settings,
                key_pool=key_pool,
                firecrawl_client=firecrawl_client,
            )
            self._firecrawl_local.service = svc
        return svc

    def _reset_email_worker_state(self) -> None:
        service = getattr(self._firecrawl_local, "service", None)
        if service is not None:
            try:
                service.close()
            except Exception:  # noqa: BLE001
                pass
            delattr(self._firecrawl_local, "service")
        if hasattr(self._firecrawl_local, "rule_email_extractor"):
            delattr(self._firecrawl_local, "rule_email_extractor")
        ch_client = getattr(self._companies_house_local, "client", None)
        if ch_client is not None:
            try:
                ch_client.close()
            except Exception:  # noqa: BLE001
                pass
            delattr(self._companies_house_local, "client")

    def _get_rule_email_extractor(self) -> EnglandRuleEmailExtractor:
        extractor = getattr(self._firecrawl_local, "rule_email_extractor", None)
        if extractor is None:
            extractor = EnglandRuleEmailExtractor(self._get_firecrawl_service())
            self._firecrawl_local.rule_email_extractor = extractor
        return extractor

    def _get_companies_house_client(self) -> CompaniesHouseClient:
        client = getattr(self._companies_house_local, "client", None)
        if client is None:
            client = CompaniesHouseClient()
            self._companies_house_local.client = client
        return client


def run_companyname_pipeline(
    *,
    config: CompanyNameConfig,
    skip_gmap: bool = False,
    skip_firecrawl: bool = False,
) -> None:
    runner = CompanyNamePipelineRunner(
        config=config,
        skip_gmap=skip_gmap,
        skip_firecrawl=skip_firecrawl,
    )
    runner.run()
